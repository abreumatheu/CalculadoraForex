import tkinter as tk
from tkinter import messagebox
from forex_python.converter import CurrencyRates
from datetime import datetime
import openpyxl
import os
from tkcalendar import DateEntry  # Importa o widget de calendário

def obter_cotacao_dolar(data, tipo='venda'):
    c = CurrencyRates()
    data_formatada = datetime.strptime(data, '%Y-%m-%d')

    if tipo == 'venda':
        cotacao_dolar = c.get_rate('USD', 'BRL', data_formatada)
    elif tipo == 'compra':
        cotacao_dolar = 1 / c.get_rate('BRL', 'USD', data_formatada)
    else:
        raise ValueError("O tipo deve ser 'compra' ou 'venda'.")

    return cotacao_dolar

def calcular_imposto(data_entrada, data_saida, valor_entrada, valor_saida):
    cotacao_compra = obter_cotacao_dolar(data_entrada, tipo='compra')
    cotacao_venda = obter_cotacao_dolar(data_saida, tipo='venda')

    lucro_dolar = valor_saida - valor_entrada
    lucro_real = lucro_dolar * cotacao_venda

    taxa_imposto = 0.15
    imposto_pagar = lucro_real * taxa_imposto

    return imposto_pagar

def escrever_em_excel(valor_imposto, cpf):
    # Caminho completo para o arquivo Excel
    caminho_planilha = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'nova_planilha.xlsx')

    # Se o arquivo já existir, carrega-o, caso contrário, cria uma nova planilha
    if os.path.isfile(caminho_planilha):
        workbook = openpyxl.load_workbook(caminho_planilha)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # Adiciona um cabeçalho na primeira linha
        sheet['H1'] = 'Imposto a Pagar'
        sheet['I1'] = 'CPF'

    # Encontra a próxima linha disponível na coluna I
    nova_linha = sheet.max_row + 1

    # Adiciona o valor do imposto na coluna I e o CPF na coluna J
    sheet[f'I{nova_linha}'] = valor_imposto
    sheet[f'J{nova_linha}'] = cpf

    # Salva a planilha
    workbook.save(caminho_planilha)

def calcular_e_exibir_imposto():
    data_entrada = entry_data_entrada.get_date().strftime('%Y-%m-%d')
    data_saida = entry_data_saida.get_date().strftime('%Y-%m-%d')

    # Substituir vírgulas por pontos nas entradas de valores
    valor_entrada = float(entry_valor_entrada.get().replace(',', '.'))
    valor_saida = float(entry_valor_saida.get().replace(',', '.'))

    cpf = entry_cpf.get()

    try:
        imposto_calculado = calcular_imposto(data_entrada, data_saida, valor_entrada, valor_saida)
        resultado_label.config(text=f"CPF: {cpf}, Valor do Imposto a ser Pago: R${imposto_calculado:.2f}")

        # Escrever o valor do imposto e o CPF na nova planilha
        escrever_em_excel(imposto_calculado, cpf)

        # Limpar os campos de entrada
        entry_data_entrada.set_date(datetime.today())
        entry_data_saida.set_date(datetime.today())
        entry_valor_entrada.delete(0, tk.END)
        entry_valor_saida.delete(0, tk.END)
        entry_cpf.delete(0, tk.END)

        # Exibir mensagem de sucesso
        messagebox.showinfo("Sucesso", "Cálculo do imposto realizado e salvo com sucesso!")

        # Atualizar o resultado_label
        resultado_label.config(text="")

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao calcular imposto: {str(e)}")

# Criar a janela principal
janela = tk.Tk()
janela.title("Calculadora de Imposto Forex :)")

# Adicionar widgets à janela
label_data_entrada = tk.Label(janela, text="Data de Entrada:")
label_data_entrada.grid(row=0, column=0, padx=10, pady=5)

# Substitui a Entry padrão por DateEntry para a data de entrada
entry_data_entrada = DateEntry(janela, date_pattern="yyyy-mm-dd", width=12)
entry_data_entrada.grid(row=0, column=1, padx=10, pady=5)

label_data_saida = tk.Label(janela, text="Data de Saída:")
label_data_saida.grid(row=1, column=0, padx=10, pady=5)

# Substitui a Entry padrão por DateEntry para a data de saída
entry_data_saida = DateEntry(janela, date_pattern="yyyy-mm-dd", width=12)
entry_data_saida.grid(row=1, column=1, padx=10, pady=5)

label_valor_entrada = tk.Label(janela, text="Valor de Entrada em Dólares:")
label_valor_entrada.grid(row=2, column=0, padx=10, pady=5)
entry_valor_entrada = tk.Entry(janela)
entry_valor_entrada.grid(row=2, column=1, padx=10, pady=5)

label_valor_saida = tk.Label(janela, text="Valor de Saída em Dólares:")
label_valor_saida.grid(row=3, column=0, padx=10, pady=5)
entry_valor_saida = tk.Entry(janela)
entry_valor_saida.grid(row=3, column=1, padx=10, pady=5)    

label_cpf = tk.Label(janela, text="Número do CPF:")
label_cpf.grid(row=4, column=0, padx=10, pady=5)

entry_cpf = tk.Entry(janela)
entry_cpf.grid(row=4, column=1, padx=10, pady=5)

calcular_button = tk.Button(janela, text="Calcular Imposto", command=calcular_e_exibir_imposto)
calcular_button.grid(row=5, column=0, columnspan=2, pady=10)

resultado_label = tk.Label(janela, text="")
resultado_label.grid(row=6, column=0, columnspan=2, pady=5)

# Iniciar o loop da interface gráfica
janela.mainloop()
